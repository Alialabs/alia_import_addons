# -*- coding: utf-8 -*-
####################################
#
#    Created on 6 de abril de 2018
#
#    @author:castor
#
##############################################################################
#
# 2018 ALIA Technologies
#       http://www.alialabs.com
#
# WARNING: This program as such is intended to be used by professional
# programmers who take the whole responsability of assessing all potential
# consequences resulting from its eventual inadequacies and bugs
# End users who are looking for a ready-to-use solution with commercial
# garantees and support are strongly adviced to contract a Free Software
# Service Company
#
# This program is Free Software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA  02111-1307, USA.
#
##############################################################################


import logging
import sys
import openpyxl,base64,io
from osv import fields, osv
from tools.translate import _
from pip.utils import file_contents
from datetime import datetime
from alia_base_imports.alia_base_excel_file_handler import AliaBaseExcelFileHandler

_logger = logging.getLogger(__name__)
_schema = logging.getLogger(__name__ + '.schema')




class alia_account_account_aligment(osv.osv_memory):
    _name = 'alia.account.account.aligment'
    
    _columns = {
                'name':fields.char('Name',size=128),
                'account_import_id':fields.many2one('alia.account.import.wizard'),
                'account_move_id':fields.integer("Account Move ID"),
                'account_move_line_id':fields.integer("Account Move Line ID"),
                'account_date':fields.date("Account Date"),
                'account_account':fields.char("Account number",size=10),
                'account_name':fields.char("Account name",size=30),
                'account_debit':fields.float("Account Debit"),
                'account_credit':fields.float("Account Credit"),
                'account_balance':fields.float("Account Balance"),
                'account_journal':fields.many2one('account.journal'),
                'account_period':fields.many2one('account.period'),
                'account_account_aligment_id':fields.many2one('account.account'),
                'account_move_line_partner_id':fields.many2one('res.partner'),
                'account_move_line_concept':fields.char('Account move line concept',size=128),
                'account_move_concept':fields.char('Account move concept',size=128),
                }
    

class alia_account_account_match(osv.osv_memory):
    _name = 'alia.account.account.match'
    
    _columns = {
                'name':fields.char('Name',size=128),
                'account_import_id':fields.many2one('alia.account.import.wizard'),
                'account_original_code':fields.char("Account Original Number",size=10),
                'account_name':fields.char("Account name",size=30),
                'account_account_match_id':fields.many2one('account.account',"Account matching"),
                }   


class alia_account_import_wizard(osv.osv_memory):
    _name = 'alia.account.import.wizard'

    _columns = {
               'name':fields.char('Name',size=128),
               'file_to_import':fields.binary('File to import',required=True),
               'omit_last_rows':fields.integer('Last rows to be omitted'),
               'omit_init_rows':fields.integer('Init rows to be omitted'),
               'account_move_confirm':fields.boolean('Automatic account move confirm'),
               'one_step_wizard':fields.boolean('Wizard without checking, directly creation'),
               'override_account_name':fields.boolean('Override Account Names'),
               'fixed_account_code_len':fields.boolean('Using fixed length for accounting matching'),
               'automatic_account_creation':fields.boolean('Automatic account creation in case of aligment not found'),
               'default_journal_id':fields.many2one('account.journal','Default Journal for account moves',required=True),
               'default_period_id': fields.many2one('account.period', 'Default Period', required=True),
               'company_id': fields.many2one('res.company', 'Company', required=True),
               'account_aligments':fields.one2many('alia.account.account.aligment','account_import_id'),
               'name_filter_opening_account_move':fields.char('Text filter for opening account move',size=40),
               'name_filter_closing_account_move':fields.char('Text filter for closing account move',size=40),
               'name_filter_results_account_move':fields.char('Text filter for results account move',size=40),
               'account_matches':fields.one2many('alia.account.account.match','account_import_id'),
               'account_creation_on_fly':fields.boolean('Account creation on the fly. ¡HIGH RISK ON THIS PROCEDURE!'),
               'account_automatic_params_search':fields.boolean('Automatic search for account params in creation process'),
               'state':fields.selection([('preprocess','preprocess'),('aligment','aligment')])
               }
    
    _defaults = {
                 'name':"Wizard Import",
                 'state':'preprocess',
                 'name_filter_opening_account_move':'Open Account Move Ref',
                 'name_filter_closing_account_move':'Close Account Move Ref',
                 'name_filter_results_account_move':'Balance Account Move Ref',
                 'fixed_account_code_len':True,
                 'automatic_account_creation':True,
                 'override_account_name':False,
                 'account_move_confirm':True,
                 }
    
    

    def _locate_most_significant_code_value(self,cr,uid,account_number_code,context=None):
        """
        This method locates the most significant code value from account code.
        I.e.  4325000 -> 4325    i.e. 4325001 -> 4325001
        """
        code = account_number_code[::-1]
        for c in code:
            if c == '0':
                code = code[1:]
            else:
                break;
        return code[::-1]
    
    

    def _locate_account_move_period(self,cr,uid,ids,wzd,name,date,company_id,context=None):
        """
        
        """
        account_period_obj = self.pool.get('account.period')
        wizard_main_obj = self.browse(cr, uid, ids[0])
        #Special case: opening period
        if wizard_main_obj.name_filter_opening_account_move == name or wizard_main_obj.name_filter_closing_account_move == name:
            ids = account_period_obj.search(cr,uid,[('date_start','=',date),('date_stop','=',date),('company_id','=',company_id.id),('special','=',True)],context=context)
            return ids[0] if ids else wzd.default_period_id.id
        #Special case: results period
        if wizard_main_obj.name_filter_results_account_move == name:
            ids = account_period_obj.search(cr,uid,[('date_start','<=',date),('date_stop','>=',date),('company_id','=',company_id.id),('special','=',True)],context=context)
            return ids[0] if ids else wzd.default_period_id.id
        #Normal case: locate the proper period    
        ids = account_period_obj.search(cr,uid,[('date_start','<=',date),('date_stop','>=',date),('company_id','=',company_id.id),('special','=',False)],context=context)
        if not ids:
            raise ValueError("Period not found! Be sure that properly account periods are created!!") 
        
        return ids[0] if ids else wzd.default_period_id.id
   
   
    def _find_parent_account_id(self, cr, uid, wzd, account_code, context=None):
        """
        Finds the parent account given an account code.
        It will remove the last digit of the code until it finds an account that
        matches exactly the code.
        """
        if len(account_code) > 0:
            parent_account_code = account_code[:-1]
            while len(parent_account_code) > 0:
                account_ids = self.pool.get('account.account').search(cr, uid, [
                                    ('code', '=', parent_account_code),
                                    ('company_id', '=', wzd.company_id.id)
                                ])
                if account_ids and len(account_ids) > 0:
                    return account_ids[0]
                parent_account_code = parent_account_code[:-1]
        # No parent found
        return None


    def _find_brother_account_id(self, cr, uid, wzd, account_code, context=None):
        """
        Finds a brother account given an account code.
        It will remove the last digit of the code until it finds an account that
        matches the begin of the code.
        """
        if len(account_code) > 0:
            brother_account_code = account_code[:-1]
            while len(brother_account_code) > 0:
                account_ids = self.pool.get('account.account').search(cr, uid, [
                                    ('code', '=like', brother_account_code + '%%'),
                                    ('company_id', '=', wzd.company_id.id)
                                ])
                if account_ids and len(account_ids) > 0:
                    return account_ids[0]
                brother_account_code = brother_account_code[:-1]
        # No brother found
        return None
    
    
    def _find_brother_account_type_not_view(self, cr, uid, wzd, account_code, context=None):
        """
        Finds a brother account given an account code.
        It will remove the last digit of the code until it finds an account that
        matches the begin of the code.
        """
        if len(account_code) > 0:
            brother_account_code = account_code[:-1]
            while len(brother_account_code) > 0:
                account_ids = self.pool.get('account.account').search(cr, uid, [
                                    ('code', '=like', brother_account_code + '%%'),
                                    ('company_id', '=', wzd.company_id.id),
                                    ('type','!=','view')
                                ])
                if account_ids and len(account_ids) > 0:
                    return self.pool.get('account.account').browse(cr,uid,account_ids[0],context=context).type
                brother_account_code = brother_account_code[:-1]
        # No brother found
        return 'other'

       
    def _create_account(self,cr,uid,wzd,account_number_code,account_name,context=None):
        """
        
        """
        _logger.info("_create_account %s with name %s",str(account_number_code),account_name)
        parent_account_id = self._find_parent_account_id(cr, uid, wzd, account_number_code)
        parent_account = self.pool.get('account.account').browse(cr, uid, parent_account_id)
        account_vals = {}
        
        if wzd.account_automatic_params_search:
            brother_account_id = self._find_brother_account_id(cr, uid, wzd, account_number_code)
            brother_account = self.pool.get('account.account').browse(cr, uid, brother_account_id)
            account_type = brother_account.type if brother_account.type != 'view' else 'other'
            account_vals = {
                    'code': account_number_code,
                    'name': account_name,
                    'parent_id': parent_account_id,
                    'type': account_type,
                    'user_type': brother_account.user_type.id,
                    'reconcile': brother_account.reconcile,
                    'company_id': wzd.company_id.id,
                    'currency_id': brother_account.currency_id.id,
                    'currency_mode': brother_account.currency_mode,
                    'active': 1,
                    'tax_ids': [(6, 0, [tax.id for tax in brother_account.tax_ids])],
                    'note': False,
                }
        else:
            account_vals = {
                    'code': account_number_code,
                    'name': account_name,
                    'parent_id': parent_account_id,
                    'type': 'other',
                    'user_type': parent_account.user_type.id,
                    'reconcile': True,
                    'company_id': wzd.company_id.id,
                    'currency_mode': parent_account.currency_mode,
                    'active': 1,
                    'note': False,
                }
            
        
        if parent_account_id:                  
            account_id = self.pool.get('account.account').create(cr, uid, account_vals)
            if wzd.account_creation_on_fly:
                cr.commit()
        _logger.info("Account Created %s",str(account_id))
        return account_id if account_id else False
    
    
    def _locate_res_partner_by_reference(self,cr,uid,ids,wzd,ref,context=None):
        partner_obj = self.pool.get('res.partner')
        for p in partner_obj.search(cr,uid,[('ref','=',ref)],context=context):
            return p

    def _locate_account_aligment(self,cr,uid,wzd,account_number_code, account_name, context=None):
        """
        Algorithm to locate the best choice for account aligment. This algorithm must be modified to add new
        account restrictions.
        """
        _logger.info("_locate_account_aligment %s",str(account_number_code))
        account_obj = self.pool.get('account.account')
        #1 Option: Find out a complete match, i.e. xxxxxxx = xxxxxxx
        account_id = account_obj.search(cr,uid,[('code','=',account_number_code),('company_id','=',context['company_id'].id)],context=context)
        if account_id:
            _logger.info("Account %s located (ID: %s)",str(account_number_code),str(account_id))
            for account in account_obj.browse(cr,uid,account_id,context=context):
                account_override_params = {}
                #Sometimes exists exactly the same account (same code and company) but it's view.
                #In this cases we must change type, cos it's imposible create new account with same code and company.
                if account.type == 'view':
                    account_override_params['type'] = self._find_brother_account_type_not_view(cr,uid,wzd,account_number_code,context=context)
                    _logger.info("Account %s it's a type view. Overriding type by %s)",str(account_number_code),str(account_override_params['type']))
                if context['override_account_name']:
                    _logger.info("Overriding account name %s on account %s)",account_name,str(account_number_code))
                    account_override_params['name'] = account_name
                if account_override_params:
                    _logger.info("Overriding params in account %s)",str(account_number_code))
                    account_obj.write(cr,uid,[account.id],account_override_params,context=context)
                return account.id
        # If there's not a complete match and fixed lenght is not a restriction
        if not context['fixed_account_code_len']:
            #2 Option: Find out an account code which contains the original code. Less weight digit must be 0
            #          I.e. xxxxxxx in xxxxxxxy where y = 0
            account_id = account_obj.search(cr,uid,[('code','like',account_number_code),('type','!=','view'),('company_id','=',context['company_id'].id)],context=context)
            if account_id:
                code = account_obj.browse(cr,uid,account_id[0],context=context).code
                if code[0] == account_number_code[0] and code[len(code)-1] == '0':
                    return account_id[0]
                    
            #3 Option: using a most_significant_code, try to locate an account_code coincidence
            most_significant_code = self._locate_most_significant_code_value(cr,uid,account_number_code,context=context)
            account_id = account_obj.search(cr,uid,[('code','=',most_significant_code),('type','!=','view'),('company_id','=',context['company_id'].id)],context=context)
            if not account_id: #If there's not a match, finding out an appropiate candidate
                account_id = account_obj.search(cr,uid,[('code','like',most_significant_code)],context=context)
        
        if not context['automatic_account_creation']:
            return account_id[0] if account_id else False
        
        return account_id[0] if account_id else self._create_account(cr,uid,wzd,account_number_code,account_name,context=context)
    
    
    def process_file(self,cr,uid,ids,context=None):
        """
        Description...
        """
        ir_model_data = self.pool['ir.model.data']
        model_data_id = ir_model_data._get_id(cr, uid, 'alia_accounting_import', 'alia_accounting_import_precheck_wizard_view')
        res_id = ir_model_data.browse(cr, uid, model_data_id, context=context).res_id
        try:
            wizard_main_obj = self.browse(cr, uid, ids[0])
            xls_file = AliaBaseExcelFileHandler(wizard_main_obj.file_to_import)
            xls_file.load_workbook(True)
            for sheet in xls_file.get_sheets():
                max_column = sheet.max_column
                max_rows = sheet.max_row - wizard_main_obj.omit_last_rows
                init_row_range = 1 + wizard_main_obj.omit_init_rows
                vals = []
                matchs = []
                account_control_dict = []
                account_move_id = 0
                for row in sheet.iter_rows(min_row=init_row_range,max_col=max_column,max_row=max_rows):
                    row_dict = {}
                    match_dict = {}
                    wizard_context = {}
                    # Sanity-checks
                    assert len(row)>=9, _("Excel format Error")
                    if row[3].value == None: #Discard intermedia empty lines
                        continue
                    #datetime.strptime(str(row[2].value), '%d/%m/%y')
                    
                    row_date = row[2].value #Me cansan las fechas.
                    #row_date = row_date.strftime("%Y-%m-%d")
                    
                    #To identify the account_move index                    
                    if row[0].value != None and row[0].value > 0 and row[0].value != account_move_id:
                        account_move_id = row[0].value
                             
                    row_dict['account_import_id'] = ids[0]
                    row_dict['account_move_id'] = account_move_id
                    row_dict['account_move_line_id'] = row[1].value
                    row_dict['account_date'] = row_date  
                    row_dict['account_account'] = row[3].value
                    row_dict['account_name'] = row[4].value
                    row_dict['account_move_line_concept'] = row[5].value
                    row_dict['account_move_concept'] = row[10].value
                    if row[6].value < 0:
                        row_dict['account_debit'] = 0.0
                        row_dict['account_credit'] = abs(row[6].value)
                    elif row[7].value < 0:
                        row_dict['account_debit'] = abs(row[7].value)
                        row_dict['account_credit'] = 0.0
                    else:                    
                        row_dict['account_debit'] = abs(row[6].value)
                        row_dict['account_credit'] = abs(row[7].value)
                    row_dict['account_balance'] = abs(row[6].value) - abs(row[7].value)
                    row_dict['account_journal'] = wizard_main_obj.default_journal_id.id
                    row_dict['account_period'] = self._locate_account_move_period(cr,uid,ids,wizard_main_obj,row[5].value,row_date,wizard_main_obj.company_id,context=context)
                    row_dict['account_move_line_partner_id'] = self._locate_res_partner_by_reference(cr,uid,ids,wizard_main_obj,row[9].value,context=context)
                    vals.append((0,0,row_dict))                                        
            
                    if row[3].value not in account_control_dict:
                        wizard_context['company_id'] = wizard_main_obj.company_id
                        wizard_context['override_account_name'] = wizard_main_obj.override_account_name
                        wizard_context['fixed_account_code_len'] = wizard_main_obj.fixed_account_code_len
                        wizard_context['automatic_account_creation'] = wizard_main_obj.automatic_account_creation
                        account_control_dict.append(row[3].value)
                        match_dict['account_original_code'] = row[3].value
                        match_dict['account_name'] = row[4].value
                        match_dict['account_account_match_id'] = self._locate_account_aligment(cr, uid, wizard_main_obj, row[3].value, row[4].value, wizard_context) 
                        matchs.append((0,0,match_dict))
                      
            self.write(cr,uid,ids,{'account_aligments':vals,'account_matches':matchs,'state':'aligment'},context=context)
            
            if wizard_main_obj.one_step_wizard:
                self.process_account_moves(cr, uid, [wizard_main_obj.id], context)
            else:
                return {
                    'type':'ir.actions.act_window',
                    'res_model':'alia.account.import.wizard',
                    'view_id':res_id,
                    'view_mode':'form',
                    'view_type':'form',
                    'res_id':ids[0],
                    'views':[(res_id,'form')],
                    'target':'new',
                    } 
        except Exception as e:
            _logger.error(e)
            raise
        
 
    def process_account_moves(self,cr,uid,ids,context=None):
        """
        This method process the aligment between original accounts and aligment accounts, adding this information to the structre
        to create account movements (account_move)
        """
        ir_model_data = self.pool['ir.model.data']
        model_data_id = ir_model_data._get_id(cr, uid, 'alia_accounting_import', 'alia_accounting_import_process_wizard_view')
        res_id = ir_model_data.browse(cr, uid, model_data_id, context=context).res_id
        account_aligment_obj = self.pool.get('alia.account.account.aligment')
        wzd = self.pool.get('alia.account.import.wizard').browse(cr,uid,ids[0],context=context)               
        for account_match in wzd.account_matches:
            _logger.info("Find out account move lines with original account code: %s",str(account_match.account_original_code))
            ids_to_update = account_aligment_obj.search(cr,uid,[('account_account','=',account_match.account_original_code)],context=context)
            account_aligment_obj.write(cr,uid,ids_to_update,{'account_account_aligment_id':account_match.account_account_match_id.id},context=context)         
        
        if wzd.one_step_wizard:
            self.create_account_moves(cr, uid, [wzd.id], context)
        
        return {
                'type':'ir.actions.act_window',
                'res_model':'alia.account.import.wizard',
                'view_id':res_id,
                'view_mode':'form',
                'view_type':'form',
                'res_id':ids[0],
                'views':[(res_id,'form')],
                'target':'new',
                }

  
    def create_account_moves(self,cr,uid,ids,context=None):
        """
        Description 
        """
        account_move_aux_id = False
        account_move_id = False
        wzd = self.pool.get('alia.account.import.wizard').browse(cr,uid,ids[0],context=context)
        account_moves_to_confirm = []
        for account_aligment in wzd.account_aligments:
            _logger.info("Create account move (%d) line %d",account_aligment.account_move_id,account_aligment.account_move_line_id)
            account_move_vals = {}
            account_move_line_vals = {}
            if account_move_aux_id != account_aligment.account_move_id:
                account_move_aux_id = account_aligment.account_move_id
                account_move_vals['name'] = account_aligment.account_move_id
                account_move_vals['date'] = account_aligment.account_date
                account_move_vals['ref'] = account_aligment.account_aligment.account_move_concept
                account_move_vals['state'] = "draft"
                account_move_vals['period_id'] = account_aligment.account_period.id
                account_move_vals['journal_id'] = account_aligment.account_journal.id
                account_move_vals['company_id'] = wzd.company_id.id
                account_move_id = self.pool.get('account.move').create(cr,uid,account_move_vals,context=context)
                _logger.info("Account Move Created: %d",account_move_id)
                if wzd.account_move_confirm:
                    account_moves_to_confirm.append(account_move_id)
             
            account_move_line_vals['account_id'] = account_aligment.account_account_aligment_id.id
            account_move_line_vals['date'] = account_aligment.account_date
            account_move_line_vals['move_id'] = account_move_id
            account_move_line_vals['company_id'] = wzd.company_id.id
            account_move_line_vals['name'] = account_aligment.account_move_line_concept
            account_move_line_vals['journal_id'] = account_aligment.account_journal.id
            account_move_line_vals['partner_id'] = account_aligment.account_move_line_partner_id.id
            account_move_line_vals['period_id'] = account_aligment.account_period.id
            account_move_line_vals['credit'] = account_aligment.account_credit
            account_move_line_vals['debit'] = account_aligment.account_debit
            self.pool.get('account.move').write(cr,uid,[account_move_id],{'line_id':[(0,0,account_move_line_vals)]},context=context)
         
        #Check for automatic account move confirm
        if wzd.account_move_confirm:
            _logger.info("Confirm Account Moves ...")
            for account_id in account_moves_to_confirm:
                self.pool.get('account.move').button_validate(cr,uid,[account_id],context=context)
         
        return {
                'type':'ir.actions.act_window_close',
        }
        
# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4: