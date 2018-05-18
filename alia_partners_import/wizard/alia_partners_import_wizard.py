# -*- coding: utf-8 -*-
####################################
#
#    Created on 6 de abril de 2018
#
#    @author:castor
#
##############################################################################
#
# 2016 ALIA Technologies
#       http://www.alialabs.com
#
# WARNING: This program as such is intended to be used by professional
# programmers who take the whole responsability of assessing all potential
# consequences resulting from its eventual inadequacies and bugs
# End users who are looking for a ready-to-use solution with commercial
# garantees and support are strongly adviced to contract a Free Software
# Service Company
#
# This program is Free Software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA  02111-1307, USA.
#
##############################################################################


import logging
import sys
import openpyxl,base64,io
from osv import fields, osv
from tools.translate import _
from pip.utils import file_contents
from datetime import datetime
from alia_partners_import.alia_excel_formats_handler import AliaExcelFormatsHandler
from alia_base_imports.alia_base_excel_file_handler import AliaBaseExcelFileHandler


_logger = logging.getLogger(__name__)
_schema = logging.getLogger(__name__ + '.schema')




class alia_partner_to_import(osv.osv_memory):
    _name = 'alia.partner.to.import'
    
    _columns = {
                'name':fields.char('Name',size=128),
                'street':fields.char('Street',size=128),
                'vat':fields.char('VAT',size=128),
                'zip':fields.char('ZIP',size=128),
                'ref':fields.char('Reference',size=128),
                'city':fields.char('City',size=128),
                'customer':fields.boolean('customer'),
                'supplier':fields.boolean('supplier'),
                'company_id':fields.many2one('res.partner'),
                'wizard_id':fields.many2one('alia.partners.import.wizard'),
                }
 


class alia_partners_import_wizard(osv.osv_memory):
    _name = 'alia.partners.import.wizard'

    FORMATS = AliaExcelFormatsHandler().get_excel_supported_formats_list()

    _columns = {
               'name':fields.char('Name',size=128),
               'file_to_import':fields.binary('File to import',required=True),
               'omit_last_rows':fields.integer('Last rows to be omitted'),
               'omit_init_rows':fields.integer('Init rows to be omitted'),
               'one_step_wizard':fields.boolean('Wizard without checking, directly creation'),
               'company_id': fields.many2one('res.company', 'Company', required=True),
               'vat_prefix':fields.char('VAT prefix',size=128),
               'omit_incorrect_vat':fields.boolean('Omit incorrect VATs'),
               'omit_creation_if_exists_reference':fields.boolean('Omit Creation if exists reference'),
               'type_address': fields.selection([('default','default'),('invoice','invoice'),('delivery','delivery'),('other','other')],'Default Type Address',select="1"),
               'importation_type': fields.selection(FORMATS,'Format type',select="1"),
               'partners_list': fields.one2many('alia.partner.to.import','wizard_id'),
               'state':fields.selection([('preprocess','preprocess'),('aligment','aligment')])
               }
    
    _defaults = {
                 'name':"Wizard Import",
                 'state':'preprocess',
                 'type_address':'default',
                 'importation_type':'standard',
                 }
    
    
    
    
    def get_vat_number(self,cr,uid,wzd,vat,context=None):
        """
        Description...
        """
        if not vat:
            return False
        toret = str(wzd.vat_prefix)+str(vat)
        vat_country, vat_number = toret[:2].lower(), toret[2:].replace(' ', '')
        if wzd.omit_incorrect_vat:
            return toret if self.pool.get('res.partner').simple_vat_check(cr,uid,vat_country,vat_number,context=context) else False
        else:
            return toret
  
    
    def process_file(self,cr,uid,ids,context=None):
        """
        Description...
        """
        ir_model_data = self.pool['ir.model.data']
        model_data_id = ir_model_data._get_id(cr, uid, 'alia_partners_import', 'alia_partners_import_process_wizard_view')
        res_id = ir_model_data.browse(cr, uid, model_data_id, context=context).res_id
        try:
            wizard_main_obj = self.browse(cr, uid, ids[0])
            xls_file = AliaBaseExcelFileHandler(wizard_main_obj.file_to_import)
            xls_file.load_workbook(True)
            for sheet in xls_file.get_sheets():
                max_column = sheet.max_column
                max_rows = sheet.max_row - wizard_main_obj.omit_last_rows
                init_row_range = 1 + wizard_main_obj.omit_init_rows
                vals = []
                for row in sheet.iter_rows(min_row=init_row_range,max_col=max_column,max_row=max_rows):
                    index_handler = AliaExcelFormatsHandler()
                    row_dict = index_handler.get_row(wizard_main_obj.importation_type,row)
                    row_dict['vat'] = self.get_vat_number(cr,uid,wizard_main_obj,row_dict['vat'],context)
                    row_dict['company_id'] = wizard_main_obj.company_id.id
                    vals.append((0,0,row_dict))                                                  
            self.write(cr,uid,ids,{'partners_list':vals,'state':'aligment'},context=context)
            
            if wizard_main_obj.one_step_wizard:
                self.create_partners(cr, uid, [wizard_main_obj.id], context)
            else:
                return {
                    'type':'ir.actions.act_window',
                    'res_model':'alia.partners.import.wizard',
                    'view_id':res_id,
                    'view_mode':'form',
                    'view_type':'form',
                    'res_id':ids[0],
                    'views':[(res_id,'form')],
                    'target':'new',
                    } 
        except Exception as e:
            _logger.error(e)
            raise
        
 
        
    def create_partners(self,cr,uid,ids,context=None):
        """
        Description 
        """
        wzd = self.pool.get('alia.partners.import.wizard').browse(cr,uid,ids[0],context=context)
        for partner in wzd.partners_list:
            vals_partner = {}
            vals_address = {}
            vals_address['street'] = partner.street
            vals_address['zip'] = partner.zip
            vals_address['city'] = partner.city
            vals_address['type'] = 'default'  
            vals_partner['name'] = partner.name
            vals_partner['vat'] = partner.vat
            vals_partner['ref'] = partner.ref
            vals_partner['customer'] = partner.customer
            vals_partner['supplier'] = partner.supplier
            vals_partner['company_id'] = partner.company_id
            vals_partner['address'] = [(0,0,vals_address)]
            if self.omit_creation_if_exists_reference:
                if not self.pool.get('res.partner').search(cr,uid,[('ref','=',partner.ref)],context=context):
                    partner_id = self.pool.get('res.partner').create(cr,uid,vals_partner,context=context)
                    _logger.info("Partner Created: %d",partner_id)
                else:
                    _logger.info("Partner with reference %s already exists.",partner.ref)
                    
            else:
                partner_id = self.pool.get('res.partner').create(cr,uid,vals_partner,context=context)
                _logger.info("Partner Created: %d",partner_id)         
        return {
                'type':'ir.actions.act_window_close',
        }
        
# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4: