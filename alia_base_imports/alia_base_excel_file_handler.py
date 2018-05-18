# -*- coding: utf-8 -*-
#
#    Created on 9/05/18
#
#    @authors:loxo,cástor
#
#
# 2018 ALIA Technologies
#       http://www.alialabs.com
#
# WARNING: This program as such is intended to be used by professional
# programmers who take the whole responsability of assessing all potential
# consequences resulting from its eventual inadequacies and bugs
# End users who are looking for a ready-to-use solution with commercial
# garantees and support are strongly adviced to contract a Free Software
# Service Company
#
# This program is Free Software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; either version 2
# of the License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA  02111-1307, USA.
#
import openpyxl as openpyxl
import base64,io
from openpyxl.utils import get_column_letter
from openerp.tools.translate import _
from openerp.exceptions import Warning


class AliaBaseExcelFileHandler:
    """
    Excel Read, Load and Save tools
    """
    _sheets_values = None  # type: dict

    def __init__(self, filepath=False):
        """
        :param filepath: string - book file path
        """    
        self._filepath = filepath
        self._work_book = False
        self._sheet_names = False
        self._sheets_values = False
        self.load_workbook(True)
        

    def load_workbook(self, decode64, book_file_name=False):
        """
        Load workbook
        :param book_file_name: string - book file path
        :return:
        """
        file = book_file_name or self._filepath
        if decode64:
            decoded_data = base64.b64decode(file)
            file = io.BytesIO(decoded_data)
        try:
            self._work_book = openpyxl.load_workbook(file) 
            self._sheet_names = self._work_book.sheetnames
        except IOError as e:
            raise Warning("Not Workbook defined in EXCEL file. Probably you're using an incorrect format. Be sure it's an .xlsx file.")

    def _pos_process_row_dict(self, row_dict):
        """
        Hook to pos process row dict
        :param row_dict: dictionary {'row_ref': {row}}
        :return: dictionary {'row_ref': {row}}
        """
        return row_dict

    def get_row_dict_by_column_letter(self, row, sheet, max_column):
        """

        :param row:
        :return:
        """
        row_dict = {}
        for c in range(1, max_column):
            column_letter = get_column_letter(c)
            row_dict[column_letter] = sheet[column_letter + str(row)].value
        row_dict = self._pos_process_row_dict(row_dict)
        return row_dict

    def get_row_dict_by_title(self, row, sheet, max_column, row_title=1):
        """

        :param row:
        :return:
        """
        row_dict = {}
        for c in range(1, max_column):
            column_letter = get_column_letter(c)
            row_dict[sheet[column_letter + str(row_title)].value] = sheet[column_letter + str(row)].value
        row_dict = self._pos_process_row_dict(row_dict)
        return row_dict

    def get_sheets(self):
        """
        Return a sheets list
        """
        return self._work_book._sheets


    def get_sheet_dict(self, sheet_name, row_ref=False):
        """
        Return a dictionary with sheet rows as dictionary with row row_ref column as row key
        :param sheet_name: string - sheet name
        :param row_ref: string - reference as key to identify rows in sheet dictionary.
        :return: dictionary {'row_ref': {row}}
        """
        sheet = self._work_book[sheet_name]
        max_column = sheet.max_column
        max_row = sheet.max_row + 1

        if not row_ref:
            row_ref = sheet['A' + str(1)].value

        toret_dict = {}
        for row in range(2, max_row):
            row_dict = self.get_row_dict_by_title(row, sheet, max_column)
            toret_dict[row_dict[row_ref]] = row_dict

        return toret_dict

    def get_sheet_list(self, sheet_name, row_ref=False):
        """
        Return a dictionary with sheet rows as dictionary with row row_ref column as row key
        :param sheet_name: string - sheet name
        :param row_ref: string - reference as key to identify rows in sheet dictionary.
        :return: dictionary {'row_ref': {row}}
        """
        sheet = self._work_book[sheet_name]
        max_column = sheet.max_column
        max_row = sheet.max_row + 1

        toret = []
        for row in range(2, max_row):
            row_dict = self.get_row_dict_by_column_letter(row, sheet, max_column)
            toret.append(row_dict)

        return toret

    def _get_dynamic_sheet_load_method_and_type(self, container_type):
        """
        Return the method to load the sheets values.
        :param container_type: string - container type
        :return: tuple - (type, method call without parameters)
        """
        if container_type and container_type == 'dict':
            return {}, self.get_sheet_dict
        else:
            return [], self.get_sheet_list

    def load_sheets(self, container_type=False):
        """
        Return a dictionary of sheets
        :return: dictionary {'sheet_name': {sheet}}
        """
        get_sheet_dinamyc_method_and_type = self._get_dynamic_sheet_load_method_and_type(container_type)
        sheets_vals = {}
        for sheet_name in self._sheet_names:
            sheets_vals[sheet_name] = get_sheet_dinamyc_method_and_type[1](sheet_name)
        self._sheets_values = sheets_vals

        return sheets_vals

    def get_rows_in_list(self, sheet_data_dict):
        """
        Process a dict an return a list of values
        :param sheet_data_dict:
        :return:
        """
        list_toret = []
        for key in sheet_data_dict.keys():
            for row in sheet_data_dict[key]:
                list_toret.append(row)
        return list_toret

    @staticmethod
    def save_first_line(sheet, first_line_vals):
        """
        Save the first line in a sheet. Column Headers
        :param sheet: openpyxl sheet object
        :param first_line_vals: dictionary - {'A': value, ...}
        :return: sheet
        """
        index = 0
        for row in sheet.iter_cols(min_row=1, min_col=1, max_col=len(first_line_vals)):
            for cell in row:
                cell.value = first_line_vals.keys()[index]
                index += 1
        return sheet

    @staticmethod
    def save_rows(sheet, sheet_rows_data_list, start_range):
        """
        Save rows values on sheet
        :param sheet: openpyxl sheet object
        :param sheet_rows_data_list: list
        :param start_range: interger - index to start save rows
        :return: sheet
        """
        for row in range(start_range, len(sheet_rows_data_list) + start_range):
            for col in range(1, len(sheet_rows_data_list[row-start_range].keys())):
                sheet.cell(column=col, row=row, value=sheet_rows_data_list[row-start_range].values()[col-1])
        return sheet

    def save_data_in_workbook(self, data_sheets, filename, first_line=[]):
        """
        Save sheets values on workbook
        :param data_sheets: dictionary - {'sheet_name': [{row}, ..]}
        :param filename: string - name of the output file
        :param first_line: [] - sheet names when the first row is column header
        """
        new_wb = openpyxl.Workbook()
        sheet_index = 0
        for sheet_name in data_sheets.keys():
            new_wb.create_sheet(index=sheet_index, title=sheet_name)
            new_sheet = new_wb[sheet_name]

            start_range = 1
            if sheet_name in first_line:
                self.save_first_line(new_sheet, data_sheets[sheet_name][0])
                start_range = 2

            self.save_rows(new_sheet, data_sheets[sheet_name], start_range)

        new_wb.save(filename)


# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
